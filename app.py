import os
import torch
import io
import uuid
from flask import Flask, request, jsonify, render_template, session
from PIL import Image
from transformers import VisionEncoderDecoderModel, TrOCRProcessor
import openpyxl
from openpyxl import Workbook, load_workbook
from evaluate import load
from spellchecker import SpellChecker  # Import the new spell checker
import boto3
from botocore.exceptions import NoCredentialsError
from pathlib import Path
from dotenv import load_dotenv
import os

# Flask app setup
app = Flask(__name__)
app.secret_key = 'supersecretkey'  # Necessary for session management

@app.before_request
def ensure_user_session():
    if 'user_id' not in session:
        session['user_id'] = str(uuid.uuid4())  # Generate a unique ID for the session

# Set a file size limit of 16 MB for uploads
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Define checkpoint path
CHECKPOINT_PATH = "checkpoint-4200"

# Save the pre-trained configuration to the checkpoint directory if config.json is missing
if not os.path.exists(os.path.join(CHECKPOINT_PATH, "config.json")):
    print("Saving configuration from pre-trained model to the checkpoint directory.")
    pretrained_model = VisionEncoderDecoderModel.from_pretrained("microsoft/trocr-base-handwritten")
    pretrained_model.config.save_pretrained(CHECKPOINT_PATH)

# Load the processor and the fine-tuned model from the checkpoint
processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten")
model = VisionEncoderDecoderModel.from_pretrained(CHECKPOINT_PATH)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# Load CER metric for evaluation
cer_metric = load("cer", trust_remote_code=True)

# Initialize spell checker from `spellchecker.py`
spell_checker = SpellChecker(Path('data/vocab.txt'))  # Ensure the vocab file is available


# Load environment variables from .env file
load_dotenv()

s3 = boto3.client(
    's3',
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
    region_name='ap-southeast-1'
)

BUCKET_NAME = 'extrakto-storage'  # Replace with your S3 bucket name

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/extract', methods=['POST'])
def extract_text():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    ground_truth_text = request.form.get('ground_truth', '')

    try:
        # Load and preprocess the image
        img = Image.open(io.BytesIO(file.read())).convert("RGB")
        max_size = (1000, 1000)
        img.thumbnail(max_size)

        # Preprocess the image for the model
        pixel_values = processor(img, return_tensors="pt").pixel_values.to(device)

        # Generate the prediction
        output_ids = model.generate(pixel_values)
        predicted_text = processor.batch_decode(output_ids, skip_special_tokens=True)[0]

        # Run the spell checking function from `spellchecker.py`
        corrected_text = spell_checker.correct(predicted_text)

        # Calculate CER if ground truth is provided
        if ground_truth_text:
            cer = cer_metric.compute(predictions=[predicted_text], references=[ground_truth_text])
        else:
            cer = None

    except Exception as e:
        return jsonify({'error': f'Failed to process image: {str(e)}'}), 400

    return jsonify({
        'predicted_text': predicted_text,
        'corrected_text': corrected_text,
        'cer': cer
    })


# Helper function to create Excel file in memory, unique to user session
def create_excel_file(new_words):
    file_name = f"{session['user_id']}_extracted_words.xlsx"  # Use user_id to make file unique

    # Check if the Excel file already exists in S3
    try:
        existing_file = s3.get_object(Bucket=BUCKET_NAME, Key=file_name)
        file_content = existing_file['Body'].read()

        # Load the existing workbook from the S3 file
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

    except s3.exceptions.NoSuchKey:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Words"
        ws.cell(row=1, column=1, value="Extracted Text")  # Add a header

    # Find the next available row (append new words)
    next_row = ws.max_row + 1

    # Clean the words and concatenate them
    cleaned_words = [word.replace("Original:", "").replace("Correct:", "").strip() for word in new_words]
    concatenated_sentence = ' '.join(cleaned_words)

    # Append the new sentence to the next available row
    ws.cell(row=next_row, column=1, value=concatenated_sentence)

    # Save the workbook back to a BytesIO stream (in-memory)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # Move to the start of the file for reading

    return output, file_name


@app.route('/delete_excel_file', methods=['POST'])
def delete_excel_file():
    try:
        # Identify the file name based on the session user
        file_name = f"{session['user_id']}_extracted_words.xlsx"

        # Delete the file from S3
        s3.delete_object(Bucket=BUCKET_NAME, Key=file_name)

        return jsonify({'message': 'Excel file deleted successfully.'})

    except Exception as e:
        return jsonify({'error': f'Failed to delete file: {str(e)}'}), 500

# Save to S3
def upload_to_s3(excel_file, file_name):
    try:
        s3.upload_fileobj(excel_file, BUCKET_NAME, file_name)
        return True
    except Exception as e:
        print(f"Error uploading file: {e}")
        return False


# Helper function to generate pre-signed URL for S3 file
def generate_presigned_url(file_name):
    try:
        url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': BUCKET_NAME, 'Key': file_name},
            ExpiresIn=3600  # URL expires in 1 hour
        )
        return url
    except Exception as e:
        print(f"Error generating presigned URL: {e}")
        return None


# Endpoint to save words to Excel and upload to S3
@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    data = request.json
    if 'words' not in data:
        return jsonify({'error': 'No words provided'}), 400

    words = data['words']
    excel_file, file_name = create_excel_file(words)  # Create Excel file in memory

    # Upload to S3 with session-specific file name
    if upload_to_s3(excel_file, file_name):
        presigned_url = generate_presigned_url(file_name)
        if presigned_url:
            return jsonify({'download_url': presigned_url})
        else:
            return jsonify({'error': 'Failed to generate download link'}), 500
    else:
        return jsonify({'error': 'Failed to upload file to S3'}), 500



if __name__ == '__main__':
    app.run(debug=True, use_reloader=False, port=5001)
